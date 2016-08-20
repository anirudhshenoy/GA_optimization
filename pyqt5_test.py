# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'Wing_Optimizer GUI.ui'
#
# Created by: PyQt5 UI code generator 5.6
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox

import random
import numpy
import math
import openpyxl

from deap import base
from deap import creator
from deap import tools
from deap import algorithms


class Display(object):
    def __init__(self):
        import matplotlib.pyplot as plt
        self.plt = plt
        self.h = []
        self.label = []
        self.fig, self.ax = self.plt.subplots()
        self.plt.axis('equal')
        self.plt.xlabel('x')
        self.plt.ylabel('y')
        self.plt.axis((-0.1,1.1)+self.plt.axis()[2:])
        self.ax.grid(True)
    def plot(self, X, Y,label=''):
        h, = self.plt.plot(X, Y, '-', linewidth = 1)
        self.h.append(h)
        self.label.append(label)
    def show(self,titlestring):
        self.plt.suptitle(titlestring)
        self.ax.legend(self.h, self.label)
        self.plt.show()




def optimize(b_const,c_const,alpha_const,taper_const,lift_target_kgs,v):
    
    fitness_weights=[100,50,25]
    lift_target=lift_target_kgs*9.8
    drag_target=0
    rho=1.227
    
    #e1=0.9
    e=0.85
    pi=3.1415

    
    
    ui.print_results("Opening Airfoil Database")
    wb = openpyxl.load_workbook('C:\\Users\Aniru_000\Desktop\TD-1\Airfoil\s1223\\airfoil\\AllPolars\\airfoilpolars_master.xlsx')
    ui.print_results("Database Opened")
    sheet = wb.get_sheet_by_name('slopes')                #Change to index based
    ui.print_results("Importing Data")
    slopes=[]
    zero_angle=[]                                  
    cd_coeff=[]
    
    
    airfoil_names=[]
    
    
    for cellObj in sheet.columns[0]:
        airfoil_names.append(cellObj.value)
        
    for cellObj in sheet.columns[1]:
        slopes.append(cellObj.value)
    
    for cellObj in sheet.columns[2]:
        zero_angle.append(cellObj.value)
    
    
    sheet = wb.get_sheet_by_name('cd_slopes')
    airfoil_const=[0,len(slopes)-1]
    ui.print_results(str(airfoil_const[1])+ " airfoil polars imported")
    
    for i in range(0,len(slopes)):
        zero_angle[i]=-zero_angle[i]/slopes[i]
    
    
    for rowNum in range(2,len(sheet.columns[1])+2):
        a=sheet.cell(row=rowNum, column=2).value
        b=sheet.cell(row=rowNum, column=3).value
        c=sheet.cell(row=rowNum, column=4).value
        cd_coeff.append([a,b,c])
    
    def evalOneMax(individual):
        
        #Aliases        
        b_pop=individual[0]
        c_pop=individual[1]
        taper=individual[2]
        alpha_pop=individual[3]
        AF_pop=individual[4]
        cd_pop=cd_coeff[int(AF_pop)]
        
    
        #Wing Dimensions
        MAC=(2/3)*c_pop*((taper**2 + taper +1)/(taper+1))   
        wing_area=b_pop*MAC
        AR=(b_pop**2)/wing_area
    
        #Slope Correction    
        
        a_new=(slopes[int(AF_pop)]*AR)/(2+(4+AR**2)**0.5)
        
        
        #Lift Calculation
        cl=a_new*(alpha_pop-(zero_angle[int(AF_pop)]))
        lift=0.5*rho*(v**2)*wing_area*cl
    
        #Drag Calculation
        cd_i=(cl**2)/(pi*e*AR)
        cd_p=cd_pop[0]*(alpha_pop)**2 + cd_pop[1]*(alpha_pop) + cd_pop[2]                                                      
        cd=cd_p+cd_i
        drag=0.5*rho*(v**2)*wing_area*cd
    
       
        #Fitness Value Calculations                                                         #Play around with Lift_fit parameters for convergence
        lift_fit=math.exp(-((lift-lift_target)**2)/(2*35**2))   #70,7                     #Gaussian function centered around lift_constant, A controls height
        drag_fit=math.exp(-((drag-drag_target)**2)/(2*35**2))                             #stall angle characteristics  Minimize moment
        area_fit=math.exp(-((wing_area-0)**2)/(2*35**2)) 
        fit=fitness_weights[0]*lift_fit+fitness_weights[1]*drag_fit +fitness_weights[2]*area_fit 
    
        return (fit,)
    
    def checkBounds(min, max):
            def decorator(func):
                def wrapper(*args, **kargs):
                    offspring = func(*args, **kargs)
                    for curPop in offspring:
                                    curPop[4]=math.floor(math.fabs(curPop[4]))
                                    if curPop[4]>airfoil_const[1]:
                                            curPop[4]=random.randint(0,airfoil_const[1])
                                    if curPop[3]>alpha_const[1] or curPop[3]<alpha_const[0]:
                                            curPop[3]=numpy.random.uniform(alpha_const[0],alpha_const[1])
                                    if curPop[1]>c_const[1] or curPop[1]<c_const[0]:
                                            curPop[1]=numpy.random.uniform(c_const[0],c_const[1])
                                    if curPop[0]>b_const[1] or curPop[0]<b_const[0]:
                                            curPop[0]=numpy.random.uniform(b_const[0],b_const[1])
                                    if curPop[2]>taper_const[1] or curPop[2]<taper_const[0]:
                                            curPop[2]=numpy.random.uniform(taper_const[0],taper_const[1])
        
                    return offspring
                return wrapper
            return decorator 
        
    def make_wing(span,chord, taper_ratio):
        tip_chord=taper_ratio*chord
        offset=(chord-tip_chord)/2
        y_coord=[offset,tip_chord+offset,chord,0,chord,offset+tip_chord,offset,0,offset]
        x_coord=[0,0,span/2,span/2,span/2,span,span,span/2,0]
        return x_coord,y_coord
    
    def display_results(individual):
        ui.print_results("\n Optimized Wing Specs")
        ui.print_results("********* **** ****")
                #Aliases        
        b_pop=individual[0]
        c_pop=individual[1]
        taper=individual[2]
        alpha_pop=individual[3]
        AF_pop=individual[4]
        cd_pop=cd_coeff[int(AF_pop)]
        
    
        #Wing Dimensions
        MAC=(2/3)*c_pop*((taper**2 + taper +1)/(taper+1))   
        wing_area=b_pop*MAC
        AR=(b_pop**2)/wing_area
    
        #Slope Correction    
    
        
        a_new=(slopes[int(AF_pop)]*AR)/(2+(4+AR**2)**0.5)
        
    
        #Lift Calculation
        cl=a_new*(alpha_pop-(zero_angle[int(AF_pop)]))
        lift=0.5*rho*(v**2)*wing_area*cl
    
        #Drag Calculation
        cd_i=(cl**2)/(pi*e*AR)
        cd_p=cd_pop[0]*(alpha_pop)**2 + cd_pop[1]*(alpha_pop) + cd_pop[2]                                                      
        cd=cd_p+cd_i
        drag=0.5*rho*(v**2)*wing_area*cd 
        
        ui.print_results("Span: " + str(b_pop)[0:4])
        ui.print_results("Chord: "+str(c_pop)[0:4])      
        ui.print_results("Taper Ratio: "+str(taper)[0:4])
        ui.print_results("Tip Chord: "+str(taper*c_pop)[0:4])
        ui.print_results("Lift (kgs): " +str(lift/9.8)[0:5])
        ui.print_results("Airfoil: " +airfoil_names[int(AF_pop)])
        ui.print_results("Angle (degs): " +str(alpha_pop)[0:4])
        ui.print_results("Drag(kgs): " +str(drag/9.8)[0:5])
        sheet = wb.get_sheet_by_name('x')
        x_coord=[]
        for cellObj in sheet.rows[int(hof[0][4])+1]:
                x_coord.append(cellObj.value)
                
        sheet = wb.get_sheet_by_name('y') 
        y_coord=[]
        for cellObj in sheet.rows[int(hof[0][4])+1]:
                y_coord.append(cellObj.value)
                
        d=Display()
        d.plot(x_coord[1:],y_coord[1:])
        d.show('Airfoil Section')
        w=Display()
        
        w.plt.axis((-0.1,b_pop+0.1)+w.plt.axis()[2:])
        x_c,y_c=make_wing(b_pop, c_pop, taper)
        w.plot(x_c,y_c)
        w.show('Wing:Top-view') 
        
        
        
        

            
    
    
    
    
    creator.create("FitnessMax", base.Fitness,weights=(1.0,1.0,1.0 ,1.0,))
    creator.create("Individual", list,typecode='i', fitness=creator.FitnessMax)
    
    IND_SIZE=5
    
    
    toolbox = base.Toolbox()
    
    toolbox.register("attr_span", random.uniform, b_const[0], b_const[1])
    toolbox.register("attr_chord", random.uniform, c_const[0], c_const[1])
    toolbox.register("attr_taper", random.uniform, taper_const[0],taper_const[1])
    toolbox.register("attr_alpha", random.uniform, alpha_const[0],alpha_const[1])
    toolbox.register("attr_airfoil", random.randint, airfoil_const[0],airfoil_const[1])
    
    N_CYCLES=1
    toolbox.register("individual", tools.initCycle, creator.Individual,
                     (toolbox.attr_span, toolbox.attr_chord, toolbox.attr_taper, toolbox.attr_alpha,toolbox.attr_airfoil), n=N_CYCLES)
    
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    
    
    
    
    toolbox.register("evaluate", evalOneMax)
    toolbox.register("mate", tools.cxTwoPoint)
    toolbox.register("mutate", tools.mutGaussian, mu=1.0, sigma=0.2, indpb=0.05)
    toolbox.register("select", tools.selTournament, tournsize=3)
    
    toolbox.decorate("mate", checkBounds(0, 1))
    toolbox.decorate("mutate", checkBounds(0, 1))
    
    ui.print_results("Starting Optimization")
    pop = toolbox.population(n=300)
    #print(pop)
    hof = tools.HallOfFame(1)
    stats = tools.Statistics(lambda ind: ind.fitness.values)
    stats.register("avg", numpy.mean)
    stats.register("std", numpy.std)
    stats.register("min", numpy.min)
    stats.register("max", numpy.max)
    
    pop, log = algorithms.eaSimple(pop, toolbox, cxpb=0.8, halloffame=hof, mutpb=0.05, ngen=40,stats=stats, verbose=False)
    
    display_results(hof[0])
    


    
    





class Ui_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(619, 608)
        self.span_min = QtWidgets.QLineEdit(Dialog)
        self.span_min.setText("0.1")
        self.span_min.setGeometry(QtCore.QRect(140, 50, 71, 21))
        self.span_min.setObjectName("span_min")
        self.Span_label = QtWidgets.QLabel(Dialog)
        self.Span_label.setGeometry(QtCore.QRect(20, 50, 91, 21))
        self.Span_label.setObjectName("Span_label")
        self.Chord_label = QtWidgets.QLabel(Dialog)
        self.Chord_label.setGeometry(QtCore.QRect(20, 90, 91, 21))
        self.Chord_label.setObjectName("Chord_label")
        self.span_max = QtWidgets.QLineEdit(Dialog)
        self.span_max.setGeometry(QtCore.QRect(270, 50, 71, 21))
        self.span_max.setObjectName("span_max")
        self.span_max.setText("2")
        self.Chord_label_2 = QtWidgets.QLabel(Dialog)
        self.Chord_label_2.setGeometry(QtCore.QRect(20, 130, 91, 21))
        self.Chord_label_2.setObjectName("Chord_label_2")
        self.chord_min = QtWidgets.QLineEdit(Dialog)
        self.chord_min.setGeometry(QtCore.QRect(140, 90, 71, 21))
        self.chord_min.setObjectName("chord_min")
        self.chord_min.setText("0.1")
        self.chord_max = QtWidgets.QLineEdit(Dialog)
        self.chord_max.setGeometry(QtCore.QRect(270, 90, 71, 21))
        self.chord_max.setObjectName("chord_max")
        self.chord_max.setText("0.4")
        self.taper_ratio_min = QtWidgets.QLineEdit(Dialog)
        self.taper_ratio_min.setGeometry(QtCore.QRect(140, 130, 71, 21))
        self.taper_ratio_min.setObjectName("taper_ratio_min")
        self.taper_ratio_min.setText("0.5")
        self.taper_ratio_max = QtWidgets.QLineEdit(Dialog)
        self.taper_ratio_max.setGeometry(QtCore.QRect(270, 130, 71, 21))
        self.taper_ratio_max.setObjectName("taper_ratio_max")
        self.taper_ratio_max.setText("1")
        self.min_label = QtWidgets.QLabel(Dialog)
        self.min_label.setGeometry(QtCore.QRect(140, 10, 71, 21))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.min_label.setFont(font)
        self.min_label.setObjectName("min_label")
        self.max_label = QtWidgets.QLabel(Dialog)
        self.max_label.setGeometry(QtCore.QRect(270, 10, 71, 21))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.max_label.setFont(font)
        self.max_label.setObjectName("max_label")
        self.angle_min = QtWidgets.QLineEdit(Dialog)
        self.angle_min.setGeometry(QtCore.QRect(140, 170, 71, 21))
        self.angle_min.setObjectName("angle_min")
        self.angle_min.setText("0")
        self.Chord_label_3 = QtWidgets.QLabel(Dialog)
        self.Chord_label_3.setGeometry(QtCore.QRect(20, 170, 91, 21))
        self.Chord_label_3.setObjectName("Chord_label_3")
        self.angle_max = QtWidgets.QLineEdit(Dialog)
        self.angle_max.setGeometry(QtCore.QRect(270, 170, 71, 21))
        self.angle_max.setObjectName("angle_max")
        self.angle_max.setText("3")
        self.optimize_button = QtWidgets.QPushButton(Dialog)
        self.optimize_button.setGeometry(QtCore.QRect(430, 120, 101, 41))
        self.optimize_button.setObjectName("optimize_button")
        self.results = QtWidgets.QTextEdit(Dialog)
        self.results.setGeometry(QtCore.QRect(10, 330, 571, 271))
        self.results.setObjectName("results")
        self.payload_target = QtWidgets.QLineEdit(Dialog)
        self.payload_target.setGeometry(QtCore.QRect(140, 220, 71, 21))
        self.payload_target.setObjectName("payload_target")
        self.payload_target.setText("4.0")
        self.mtow_label = QtWidgets.QLabel(Dialog)
        self.mtow_label.setGeometry(QtCore.QRect(20, 210, 91, 21))
        self.mtow_label.setObjectName("mtow_label")
        self.results_label = QtWidgets.QLabel(Dialog)
        self.results_label.setGeometry(QtCore.QRect(20, 280, 111, 31))
        self.results_label.setObjectName("results_label")
        self.velocity_label = QtWidgets.QLabel(Dialog)
        self.velocity_label.setGeometry(QtCore.QRect(20, 250, 91, 21))
        self.velocity_label.setObjectName("velocity_label")
        self.velocity = QtWidgets.QLineEdit(Dialog)
        self.velocity.setGeometry(QtCore.QRect(140, 250, 71, 21))
        self.velocity.setObjectName("velocity")
        self.velocity.setText("12")

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)
        
        self.optimize_button.clicked.connect(self.btn_clk)
    
    def infoDialogue(self): ## Method to open a message box
        infoBox = QMessageBox() ##Message Box that doesn't run
        infoBox.setIcon(QMessageBox.Information)
        infoBox.setText("Bro, do you think this is some BMS Plane? ")
        infoBox.setWindowTitle("Are you serious?")
        infoBox.setStandardButtons(QMessageBox.Ok)
        infoBox.setEscapeButton(QMessageBox.Close)
        infoBox.exec_()       
         
    def btn_clk(self):
        angle_constraint=[float(self.angle_min.text()), float(self.angle_max.text())]
        span_constraint=[float(self.span_min.text()), float(self.span_max.text())]
        taper_constraint=[float(self.taper_ratio_min.text()),float(self.taper_ratio_max.text())]
        chord_constraint=[float(self.chord_min.text()),float(self.chord_max.text())]
        payload_target=float(self.payload_target.text())
        velocity_constraint=float(self.velocity.text())
        if (angle_constraint[0]<-5 or angle_constraint[1]>10 or span_constraint[0]<0.1 or span_constraint[1]>5 or taper_constraint[0]<0.3 or taper_constraint[1]>1 or chord_constraint[0]<0.1 or chord_constraint[1]>0.5 or velocity_constraint>25):
            self.infoDialogue()
        else:
            optimize(span_constraint,chord_constraint,angle_constraint,taper_constraint,payload_target,velocity_constraint)
        
    def print_results(self,mystring):
        self.results.append(mystring) #append string
        QtWidgets.QApplication.processEvents()
            
    
    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Wing Optimizer"))
        self.Span_label.setText(_translate("Dialog", "Wing Span"))
        self.Chord_label.setText(_translate("Dialog", "Chord"))
        self.Chord_label_2.setText(_translate("Dialog", "Taper Ratio"))
        self.min_label.setText(_translate("Dialog", "Minimum"))
        self.max_label.setText(_translate("Dialog", "Maximum"))
        self.Chord_label_3.setText(_translate("Dialog", "Operating Angle"))
        self.optimize_button.setText(_translate("Dialog", "Optimize"))
        self.mtow_label.setText(_translate("Dialog", "Target MToW (kgs)"))
        self.results_label.setText(_translate("Dialog", "Results: "))
        self.velocity_label.setText(_translate("Dialog", "Velocity (m/s)"))

        
    
       


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.show()
    sys.exit(app.exec_())

