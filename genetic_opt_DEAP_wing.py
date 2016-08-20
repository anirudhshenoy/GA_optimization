import random
import numpy
import math
import openpyxl

from deap import algorithms
from deap import base
from deap import creator
from deap import tools


b_const=[0.5,5]
c_const=[0.2,0.4]
taper_const=[0.5,0.95]
alpha_const=[0,3]

lift_target_kgs=3.6
lift_target=lift_target_kgs*9.8
drag_target=0
rho=1.227
v=10
e1=0.9
e=0.85
pi=3.1415

print ("Opening Airfoil Database")
wb = openpyxl.load_workbook('C:/Users/Aniru_000/Desktop/TD-1/Airfoil/s1223/airfoil/MasterPolarsFlat/airfoilpolars_master.xlsx')
print ("Opened File")
sheet = wb.get_sheet_by_name('slopes')                #Change to index based
print ("Opened Sheet")
slopes=[]
zero_angle=[]
temp=[]                                       #Please fix this shit
cd_coeff=[]


airfoil_names=[]


for cellObj in sheet.columns[0]:
    airfoil_names.append(cellObj.value)
    
for cellObj in sheet.columns[1]:
    slopes.append(cellObj.value)

for cellObj in sheet.columns[2]:
    zero_angle.append(cellObj.value)
    
for i in range(0,len(slopes)):
    zero_angle[i]=-zero_angle[i]/slopes[i]    

sheet = wb.get_sheet_by_name('cd_slopes')
airfoil_const=[0,len(slopes)-1]


for rowNum in range(2,len(sheet.columns[1])+2):
    a=sheet.cell(row=rowNum, column=2).value
    b=sheet.cell(row=rowNum, column=3).value
    c=sheet.cell(row=rowNum, column=4).value
    cd_coeff.append([a,b,c])
    





creator.create("FitnessMax", base.Fitness,weights=(1.0,1.0,1.0 ,1.0,))
creator.create("Individual", list,typecode='i', fitness=creator.FitnessMax)

IND_SIZE=5


toolbox = base.Toolbox()

toolbox.register("attr_span", random.uniform, b_const[0], b_const[1])
toolbox.register("attr_chord", random.uniform, c_const[0], c_const[1])
toolbox.register("attr_taper", random.uniform, taper_const[0],taper_const[1])
toolbox.register("attr_alpha", random.uniform, alpha_const[0],alpha_const[1])
toolbox.register("attr_airfoil", random.randint, airfoil_const[0],airfoil_const[1])

N_CYCLES=1
toolbox.register("individual", tools.initCycle, creator.Individual,
                 (toolbox.attr_span, toolbox.attr_chord, toolbox.attr_taper, toolbox.attr_alpha,toolbox.attr_airfoil), n=N_CYCLES)

toolbox.register("population", tools.initRepeat, list, toolbox.individual)

def evalOneMax(individual):
    
    #Aliases        
    b_pop=individual[0]
    c_pop=individual[1]
    taper=individual[2]
    alpha_pop=individual[3]
    AF_pop=individual[4]
    cd_pop=cd_coeff[int(AF_pop)]
    

    #Wing Dimensions
    MAC=(2/3)*c_pop*((taper**2 + taper +1)/(taper+1))   
    wing_area=b_pop*MAC
    AR=(b_pop**2)/wing_area

    #Slope Correction    
    #a_new=slopes[int(AF_pop)]/(pi*e1*AR)
    #a_new=slopes[int(AF_pop)]/(1+(57.3*a_new))
    
    a_new=(slopes[int(AF_pop)]*AR)/(2+(4+AR**2)**0.5)
    #a_new=a_new/57.29
        
        #Lift Calculation
    cl=a_new*(alpha_pop-(zero_angle[int(AF_pop)]))

    #Lift Calculation
    #cl=intercepts[int(AF_pop)]+(a_new*alpha_pop)
    lift=0.5*rho*(v**2)*wing_area*cl

    #Drag Calculation
    cd_i=(cl**2)/(pi*e*AR)
    cd_p=cd_pop[0]*(alpha_pop)**2 + cd_pop[1]*(alpha_pop) + cd_pop[2]                                                      
    cd=cd_p+cd_i
    drag=0.5*rho*(v**2)*wing_area*cd

   
    #Fitness Value Calculations                                                         #Play around with Lift_fit parameters for convergence
    lift_fit=100*math.exp(-((lift-lift_target)**2)/(2*35**2))   #70,7                     #Gaussian function centered around lift_constant, A controls height
    drag_fit=50*math.exp(-((drag-drag_target)**2)/(2*35**2))                             #stall angle characteristics  Minimize moment
    area_fit=25*math.exp(-(((wing_area-0)*10)**2)/(2*35**2))
    fit=lift_fit+drag_fit +area_fit 

    return (fit,)

def checkBounds(min, max):
    def decorator(func):
        def wrapper(*args, **kargs):
            offspring = func(*args, **kargs)
            for curPop in offspring:
                            curPop[4]=math.floor(math.fabs(curPop[4]))
                            if curPop[4]>airfoil_const[1]:
                                    curPop[4]=random.randint(0,airfoil_const[1])
                            if curPop[3]>alpha_const[1] or curPop[3]<alpha_const[0]:
                                    curPop[3]=numpy.random.uniform(alpha_const[0],alpha_const[1])
                            if curPop[1]>c_const[1] or curPop[1]<c_const[0]:
                                    curPop[1]=numpy.random.uniform(c_const[0],c_const[1])
                            if curPop[0]>b_const[1] or curPop[0]<b_const[0]:
                                    curPop[0]=numpy.random.uniform(b_const[0],b_const[1])
                            if curPop[2]>taper_const[1] or curPop[2]<taper_const[0]:
                                    curPop[2]=numpy.random.uniform(taper_const[0],taper_const[1])

            return offspring
        return wrapper
    return decorator               

toolbox.register("evaluate", evalOneMax)
toolbox.register("mate", tools.cxTwoPoint)
toolbox.register("mutate", tools.mutGaussian, mu=1.0, sigma=0.2, indpb=0.05)
toolbox.register("select", tools.selTournament, tournsize=3)

toolbox.decorate("mate", checkBounds(0, 1))
toolbox.decorate("mutate", checkBounds(0, 1))

def display_results(individual):
        print("\n Optimized Wing Specs")
        print("********* **** ****")
                #Aliases        
        b_pop=individual[0]
        c_pop=individual[1]
        taper=individual[2]
        alpha_pop=individual[3]
        AF_pop=individual[4]
        cd_pop=cd_coeff[int(AF_pop)]
        
    
        #Wing Dimensions
        MAC=(2/3)*c_pop*((taper**2 + taper +1)/(taper+1))   
        wing_area=b_pop*MAC
        AR=(b_pop**2)/wing_area
    
        #Slope Correction    
        print(slopes[int(AF_pop)])
        #a_new=slopes[int(AF_pop)]/(pi*e1*AR)
        #a_new=slopes[int(AF_pop)]/(1+(57.3*a_new))
        
        a_new=(slopes[int(AF_pop)]*AR)/(2+(4+AR**2)**0.5)
        #a_new=a_new/57.29
        print(AR)
        print(a_new)
       
        #Lift Calculation
        cl=a_new*(alpha_pop-(zero_angle[int(AF_pop)]))
        print(cl)
        lift=0.5*rho*(v**2)*wing_area*cl
    
        #Drag Calculation
        cd_i=(cl**2)/(pi*e*AR)
        cd_p=cd_pop[0]*(alpha_pop)**2 + cd_pop[1]*(alpha_pop) + cd_pop[2]                                                      
        cd=cd_p+cd_i
        drag=0.5*rho*(v**2)*wing_area*cd 
        
        print("Span: " + str(b_pop)[0:4])
        print("Chord: "+str(c_pop)[0:4])      
        print("Taper Ratio: "+str(taper)[0:4])
        print("Tip Chord: "+str(taper*c_pop)[0:4])
        print("Lift (kgs): " +str(lift/9.8)[0:5])
        print("Airfoil: " +airfoil_names[int(AF_pop)])
        print("Angle (degs): " +str(alpha_pop)[0:4]) 
        print("Cl/cd: " +str(cl/cd)[0:5]) 
    

def main():
    
    
    pop = toolbox.population(n=500)
    #print(pop)
    hof = tools.HallOfFame(1)
    stats = tools.Statistics(lambda ind: ind.fitness.values)
    stats.register("avg", numpy.mean)
    stats.register("std", numpy.std)
    stats.register("min", numpy.min)
    stats.register("max", numpy.max)
    
    pop, log = algorithms.eaSimple(pop, toolbox, cxpb=0.8, halloffame=hof, mutpb=0.10, ngen=100,stats=stats, verbose=True)
    
    display_results(hof[0])
    return pop, log, hof

if __name__ == "__main__":
    main()
