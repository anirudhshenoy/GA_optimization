import csv
import os
import openpyxl
from openpyxl.styles import Alignment
import numpy as np

airfoil_no=1
wb=openpyxl.Workbook()
lift=wb.create_sheet(index=0, title='Cl')
drag=wb.create_sheet(index=1, title='Cd')
pres=wb.create_sheet(index=2, title='Cdp')
mom=wb.create_sheet(index=3, title='Cm')
txtr=wb.create_sheet(index=4, title='top_xtr')
bxtr=wb.create_sheet(index=5, title='bot_xtr')
slopes=wb.create_sheet(index=6, title='slopes')
cd_slopes=wb.create_sheet(index=7, title='cd_slopes')
cm_slopes=wb.create_sheet(index=7, title='cm_slopes')
wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
sheets=[lift,drag,pres,mom,txtr,bxtr]

angles=np.arange(-6,10,0.5)
ids=list(range(2,len(angles)+2))
angle_keys=dict(zip(angles,ids))

    
for s in sheets:
    colCounter=0
    s.column_dimensions['A'].width = 20
    for i in angles:
        cell = s.cell(row=1,column=colCounter+2)
        cell.value=str(i)
        cell.alignment = Alignment(horizontal="center")
        cell.font = cell.font.copy(bold=True)
        colCounter+=1

os.chdir('C:\\Users\Aniru_000\Desktop\TD-1\Airfoil\s1223\\airfoil\\Tail_Polars')
#start Airfoil Loop Here
for file in os.listdir(os.getcwd()):
    if file.endswith(".txt"):
        print('Parsing File:'+file+"   Airfoil No:"+str(airfoil_no))
        
        with open(file, newline='') as f:

            csv_f=csv.reader(f)
            rowNum=0
            alpha=[]
            cl=[]
            cd=[]
            cdp=[]
            cm=[]
            top_xtr=[]
            bot_xtr=[]
            for row in csv_f:
                
                if rowNum==2:           
                    airfoil=str(row)
                    
                    airfoil=airfoil.split(':')
                    print(airfoil)
                    #airfoil_1=airfoil[1].split(' ')[1]
                    airfoil=airfoil[1][:-2]
                    print("Airfoil Name: "+airfoil)
                if rowNum>=11:
                    
                    rowData=str(row)
                    rowData=rowData.split()
                    if len(rowData)>1:
                        alpha.append(float(rowData[1]))
                        cl.append(float(rowData[2]))
                        cd.append(float(rowData[3]))
                        cdp.append(float(rowData[4]))
                        cm.append(float(rowData[5]))
                        top_xtr.append(float(rowData[6]))
                        bot_xtr.append(float(rowData[7][0:6]))
                    
                rowNum+=1

            dat={1:cl,2:cd,3:cdp,4:top_xtr,5:bot_xtr}

            

            if len(alpha)>10:
                slopes['A'+str(airfoil_no+1)].value=airfoil
                slopes['A'+str(airfoil_no+1)].alignment=Alignment(horizontal="center")
                slopes['A'+str(airfoil_no+1)].font=cell.font.copy(bold=True)
                slopes.freeze_panes='B2'
                cd_slopes['A'+str(airfoil_no+1)].value=airfoil
                cd_slopes['A'+str(airfoil_no+1)].alignment=Alignment(horizontal="center")
                cd_slopes['A'+str(airfoil_no+1)].font=cell.font.copy(bold=True)
                cd_slopes.freeze_panes='B2'
                cm_slopes['A'+str(airfoil_no+1)].value=airfoil
                cm_slopes['A'+str(airfoil_no+1)].alignment=Alignment(horizontal="center")
                cm_slopes['A'+str(airfoil_no+1)].font=cell.font.copy(bold=True)
                cm_slopes.freeze_panes='B2'
                
                z=np.polyfit(alpha,cl,1)
                a0=z[0]
                a1=z[1]
                z=np.polyfit(alpha,cd,2)
                d0=z[0]                 #Fix this crap
                d1=z[1]
                d2=z[2]                 #Add outlier rejection
                z=np.polyfit(alpha,cm,2)
                m0=z[0]
                m1=z[1]
                m2=z[2]
                
           
                slopes['B'+str(airfoil_no+1)].value=a0
                slopes['C'+str(airfoil_no+1)].value=a1
                cd_slopes['B'+str(airfoil_no+1)].value=d0
                cd_slopes['C'+str(airfoil_no+1)].value=d1
                cd_slopes['D'+str(airfoil_no+1)].value=d2
                cm_slopes['B'+str(airfoil_no+1)].value=m0
                cm_slopes['C'+str(airfoil_no+1)].value=m1
                cm_slopes['D'+str(airfoil_no+1)].value=m2
                    


            #Write to XLS     
             
                for s in sheets:
                        s['A'+str(airfoil_no+1)].value=airfoil
                        s['A'+str(airfoil_no+1)].alignment=Alignment(horizontal="center")
                        s['A'+str(airfoil_no+1)].font=cell.font.copy(bold=True)
                        s.freeze_panes='B2'

                dat={1:cl,2:cd,3:cdp,4:cm,5:top_xtr,6:bot_xtr}
                for i in range(len(alpha)):
                        col_no=angle_keys[float(alpha[i])]
                        dat_value=1
                        for s in sheets:
                            cell=s.cell(row=airfoil_no+1,column=col_no)
                            cell.value=dat[dat_value][i]
                            cell.alignment = Alignment(horizontal="center")
                            dat_value+=1
                    
        
                airfoil_no+=1
    #if airfoil_no==1:
    #    exit()

            
           
wb.save('airfoilpolars_master.xlsx')
        
 


